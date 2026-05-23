"""CSV/XLSX export helpers."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .schema import OUTPUT_COLUMNS


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
REVIEW_FILL = PatternFill(fill_type="solid", fgColor="FDE9D9")


def _autosize_worksheet(worksheet) -> None:
    preferred_widths = {
        "A": 20,
        "B": 10,
        "C": 8,
        "D": 28,
        "E": 10,
        "F": 14,
        "G": 14,
        "H": 22,
        "I": 10,
        "J": 10,
        "K": 10,
        "L": 14,
        "M": 14,
        "N": 22,
        "O": 10,
        "P": 10,
        "Q": 10,
        "R": 14,
        "S": 18,
        "T": 40,
        "U": 20,
        "V": 34,
    }
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = preferred_widths.get(column_letter, 14)


def _format_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for row in worksheet.iter_rows(min_row=2):
        if row[-1].value and str(row[-1].value).startswith("review:"):
            for cell in row:
                cell.fill = REVIEW_FILL
    _autosize_worksheet(worksheet)


def export_records(records: list[dict], outdir: Path) -> tuple[Path, Path]:
    dataframe = pd.DataFrame(records)
    dataframe = dataframe.reindex(columns=OUTPUT_COLUMNS)
    csv_path = outdir / "table_output.csv"
    xlsx_path = outdir / "table_output.xlsx"
    dataframe.to_csv(csv_path, index=False, na_rep="")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="table")
        review_only = dataframe[dataframe["confidence_flag"].fillna("").astype(str).str.startswith("review:")]
        review_only.to_excel(writer, index=False, sheet_name="review_only")
        summary = pd.DataFrame(
            [
                {"metric": "rows_total", "value": int(len(dataframe))},
                {
                    "metric": "rows_ok",
                    "value": int((dataframe["confidence_flag"].fillna("") == "ok").sum()),
                },
                {
                    "metric": "rows_review",
                    "value": int(review_only.shape[0]),
                },
            ]
        )
        summary.to_excel(writer, index=False, sheet_name="summary")

        _format_sheet(writer.book["table"])
        _format_sheet(writer.book["review_only"])
        _format_sheet(writer.book["summary"])
    return csv_path, xlsx_path
